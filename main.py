import os
from openpyxl import load_workbook

ROOT_DIR = os.getcwd()
file_excel = os.path.join(ROOT_DIR, 'rayon_one.xlsx')

wb = load_workbook(file_excel)
sheet = wb.active

rows = sheet.max_row
columns = sheet.max_column

merged = sheet.merged_cells.ranges

for i in range(1, int(rows)):
    cell = sheet.cell(row=i, column=1)

    for merged in sheet.merged_cells.ranges:
        if (cell.coordinate in merged):
            print(cell.value)